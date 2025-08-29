#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
End-to-end local pipeline (single script)

- One Google Vision pass per image (DOCUMENT_TEXT_DETECTION)
- Smart deskew (Hough + OCR pairs) with fine grid search (in-memory)
- Build slope-aware (pre) and horizontal (post) line dumps (in-memory)
- Extract all clients & PD rows per page (robust to headers/EOBS)
- Export nicely formatted Excel via ExcelGenerator

Usage:
  python ocr_pipeline.py --input "C:\\imgs" --out "results.xlsx"
  python ocr_pipeline.py --files s1.jpg s2.jpg --out results.xlsx
  python ocr_pipeline.py --input "C:\\imgs" --out results.xlsx --deskewed-only
"""

import os
import re
import io
import cv2
import math
import glob
import argparse
import numpy as np
import pandas as pd
from typing import List, Dict, Tuple, Any, Optional
from datetime import datetime

# ========= Debug switch =========
# Set to True to re-enable saving deskewed images, writing *_lines_*.txt,
# and printing progress messages.
DEBUG = False

# ---------- Google Vision ----------
from google.cloud import vision

# ---------- openpyxl helpers ----------
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# Config (tuning)
# ============================================================
PERP_TOL_FACTOR = 0.6
SEED_BAND_H     = 3.0
ALLOW_SINGLETON = True

POST_Y_TOL_FACTOR = 0.55

# ============================================================
# Vision OCR (ONE pass per image)
# ============================================================
def _open_bytes(path: str) -> bytes:
    with open(path, "rb") as f:
        return f.read()

def extract_words_and_text(image_path: str) -> Tuple[List[Dict], str]:
    client = vision.ImageAnnotatorClient()
    resp = client.document_text_detection(image=vision.Image(content=_open_bytes(image_path)))
    if resp.error.message:
        raise RuntimeError(resp.error.message)

    full_text = resp.full_text_annotation.text or ""

    words: List[Dict] = []
    for page in resp.full_text_annotation.pages:
        for block in page.blocks:
            for para in block.paragraphs:
                for word in para.words:
                    text = "".join(s.text for s in word.symbols)
                    vs = word.bounding_box.vertices
                    xs = [v.x for v in vs]; ys = [v.y for v in vs]
                    left, top = min(xs), min(ys)
                    w, h = max(xs) - left, max(ys) - top
                    cx, cy = left + w/2.0, top + h/2.0
                    words.append({"text": text, "left": left, "top": top,
                                  "w": w, "h": h, "cx": cx, "cy": cy})
    return words, full_text

# ============================================================
# Skew estimation (Hough + OCR pairs)
# ============================================================
def weighted_median(pairs: List[Tuple[float, float]]) -> float:
    if not pairs: return 0.0
    arr = sorted(pairs, key=lambda t: t[0])
    tot = sum(w for _, w in arr)
    acc = 0.0
    for v, w in arr:
        acc += w
        if acc >= tot/2.0:
            return v
    return arr[-1][0]

def estimate_skew_pairs(words: List[Dict],
                        y_band_mult: float = 2.0,
                        min_dx_mult: float = 0.8,
                        max_abs_deg: float = 15.0) -> Tuple[float,int]:
    if not words: return 0.0, 0
    widths  = [w["w"] for w in words if w["w"]>0]
    heights = [w["h"] for w in words if w["h"]>0]
    w_med = float(np.median(widths) if widths else 10.0)
    h_med = float(np.median(heights) if heights else 16.0)
    y_band = y_band_mult * h_med
    min_dx = max(4.0, min_dx_mult * w_med)

    words_sorted = sorted(words, key=lambda w: (w["cy"], w["cx"]))
    pairs: List[Tuple[float,float]] = []
    for i, wi in enumerate(words_sorted):
        best_j = None; best_dx = None
        for j in range(i+1, len(words_sorted)):
            wj = words_sorted[j]
            dy = wj["cy"] - wi["cy"]
            if dy > y_band: break
            if abs(dy) <= y_band:
                dx = wj["cx"] - wi["cx"]
                if dx <= 0 or dx < min_dx: continue
                if best_dx is None or dx < best_dx:
                    best_dx, best_j = dx, j
        if best_j is None: continue
        wj = words_sorted[best_j]
        dx = wj["cx"] - wi["cx"]; dy = wj["cy"] - wi["cy"]
        ang = math.degrees(math.atan2(dy, dx))
        if abs(ang) <= max_abs_deg:
            pairs.append((ang, max(1.0, dx)))

    if not pairs: return 0.0, 0
    vals = np.array([v for v,_ in pairs], dtype=float)
    q1, q3 = np.percentile(vals, [25,75]); iqr = q3-q1
    lo, hi = q1 - 1.5*iqr, q3 + 1.5*iqr
    trimmed = [(v,w) for v,w in pairs if lo <= v <= hi] or pairs
    return float(weighted_median(trimmed)), len(trimmed)

def estimate_skew_hough(img: np.ndarray, thr: int = 180) -> Tuple[float,int]:
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    g = cv2.GaussianBlur(g, (3,3), 0)
    edges = cv2.Canny(g, 60, 160, apertureSize=3)
    lines = cv2.HoughLines(edges, 1, np.pi/180, threshold=thr)
    if lines is None: return 0.0, 0
    angs = []
    for (rho, theta) in lines[:,0,:]:
        ang = (theta - np.pi/2.0) * 180.0/np.pi
        while ang > 45: ang -= 90
        while ang < -45: ang += 90
        angs.append(ang)
    angs = np.array(angs, dtype=float)
    med = float(np.median(angs))
    keep = angs[np.abs(angs - med) <= 10.0]
    return (float(np.median(keep)) if keep.size else med), int(angs.size)

# ============================================================
# Rotation (image + coordinates) and scoring
# ============================================================
def rotation_matrix_keep_bounds(shape_hw: Tuple[int,int], angle_deg: float) -> Tuple[np.ndarray, Tuple[int,int]]:
    h, w = shape_hw
    center = (w/2.0, h/2.0)
    M = cv2.getRotationMatrix2D(center, angle_deg, 1.0)
    cos, sin = abs(M[0,0]), abs(M[0,1])
    new_w = int(h*sin + w*cos)
    new_h = int(h*cos + w*sin)
    M[0,2] += (new_w/2) - center[0]
    M[1,2] += (new_h/2) - center[1]
    return M, (new_h, new_w)

def rotate_image_keep_bounds(img: np.ndarray, angle_deg: float, border_value=255) -> np.ndarray:
    M, (nh, nw) = rotation_matrix_keep_bounds(img.shape[:2], angle_deg)
    return cv2.warpAffine(img, M, (nw, nh),
                          flags=cv2.INTER_LINEAR,
                          borderMode=cv2.BORDER_CONSTANT,
                          borderValue=border_value)

def transform_words(words: List[Dict], shape_hw: Tuple[int,int], angle_deg: float) -> List[Dict]:
    M, _ = rotation_matrix_keep_bounds(shape_hw, angle_deg)
    out = []
    for w in words:
        x, y = (M @ np.array([w["cx"], w["cy"], 1.0])).tolist()
        ww = dict(w)
        ww["cx_rot"], ww["cy_rot"] = float(x), float(y)
        out.append(ww)
    return out

def preview_score(img: np.ndarray, deskew_angle: float) -> float:
    h, w = img.shape[:2]
    scale = 1200.0 / max(h, w)
    small = cv2.resize(img, (int(w*scale), int(h*scale)), interpolation=cv2.INTER_AREA) if scale < 1 else img
    rot = rotate_image_keep_bounds(small, deskew_angle, border_value=255)
    resid, n = estimate_skew_hough(rot, thr=140)
    return abs(resid) if n > 0 else 90.0

# ============================================================
# Slope-based clustering (pre-rotation)
# ============================================================
def line_from_points(p0, p1):
    (x0,y0),(x1,y1)=p0,p1
    dx = x1-x0
    if abs(dx) < 1e-9: return float("inf"), x0
    m = (y1-y0)/dx; b = y0 - m*x0
    return m,b

def perp_distance(m,b,x,y):
    if math.isinf(m): return abs(x-b)
    return abs(m*x - y + b) / math.sqrt(m*m + 1.0)

def refit_line(points: List[Tuple[float,float]]) -> Tuple[float,float]:
    if len(points) == 1:
        x,y = points[0]; return 0.0, y
    xs=[p[0] for p in points]; ys=[p[1] for p in points]
    xm = sum(xs)/len(xs); ym = sum(ys)/len(ys)
    num = sum((x-xm)*(y-ym) for x,y in zip(xs,ys))
    den = sum((x-xm)**2 for x in xs)
    if abs(den) < 1e-12: return float("inf"), xm
    m = num/den; b = ym - m*xm
    return m,b

def project_t(m,b,x0,y0,x,y):
    if math.isinf(m): return y - y0
    denom = math.sqrt(1+m*m)
    return ((x-x0) + m*(y-y0))/denom

def _build_line_result(words, idxs, m, b, rotated=False):
    origin_idx = min(idxs, key=lambda i: (words[i]["cx_rot"] if rotated else words[i]["cx"]))
    x0 = words[origin_idx]["cx_rot"] if rotated else words[origin_idx]["cx"]
    y0 = words[origin_idx]["cy_rot"] if rotated else words[origin_idx]["cy"]

    ordered = sorted(
        idxs,
        key=lambda i: project_t(
            m, b, x0, y0,
            words[i]["cx_rot"] if rotated else words[i]["cx"],
            words[i]["cy_rot"] if rotated else words[i]["cy"]
        )
    )
    line_words = [words[i] for i in ordered]
    text = " ".join(w["text"] for w in line_words)

    xs = [(w["cx_rot"] if rotated else w["cx"]) for w in line_words]
    ys = [(w["cy_rot"] if rotated else w["cy"]) for w in line_words]
    return {
        "text": text,
        "words": line_words,
        "slope": m,
        "center_x": float(sum(xs)/len(xs)),
        "center_y": float(sum(ys)/len(ys)),
        "count": len(line_words),
    }

def cluster_tilted_lines(words: List[Dict]) -> List[Dict]:
    if not words: return []
    hs = sorted([w["h"] for w in words if w["h"]>0])
    h_med = hs[len(hs)//2] if hs else 16.0
    perp_tol = PERP_TOL_FACTOR * h_med
    band_dy  = SEED_BAND_H * h_med

    remaining = set(range(len(words)))
    order = sorted(remaining, key=lambda i: (words[i]["cy"], words[i]["cx"]))
    lines = []

    while remaining:
        seed_idx = next(i for i in order if i in remaining)
        remaining.remove(seed_idx)
        sx, sy = words[seed_idx]["cx"], words[seed_idx]["cy"]

        cand_idxs = [j for j in remaining if abs(words[j]["cy"] - sy) <= band_dy]
        if not cand_idxs:
            if ALLOW_SINGLETON:
                m,b = refit_line([(sx,sy)])
                lines.append(_build_line_result(words, {seed_idx}, m, b))
            continue

        cand_idxs.sort(key=lambda j: abs(words[j]["cx"] - sx))
        best_inliers = None; best_mb = None
        for j in cand_idxs[:min(10, len(cand_idxs))]:
            m,b = line_from_points((sx,sy), (words[j]["cx"], words[j]["cy"]))
            inliers = {seed_idx, j}
            for k in remaining:
                xk, yk = words[k]["cx"], words[k]["cy"]
                if perp_distance(m,b,xk,yk) <= perp_tol:
                    inliers.add(k)
            if best_inliers is None or len(inliers) > len(best_inliers):
                best_inliers, best_mb = inliers, (m,b)

        m,b = best_mb
        pts = [(words[i]["cx"], words[i]["cy"]) for i in best_inliers]
        m,b = refit_line(pts)

        expanded = set(best_inliers)
        for idx in list(remaining):
            xk, yk = words[idx]["cx"], words[idx]["cy"]
            if perp_distance(m,b,xk,yk) <= perp_tol:
                expanded.add(idx)

        for idx in expanded:
            if idx in remaining:
                remaining.remove(idx)
        lines.append(_build_line_result(words, expanded, m, b))

    lines.sort(key=lambda L: L["center_y"])
    return lines

# ============================================================
# Post-rotation grouping (simple horizontal lines)
# ============================================================
def group_horizontal_lines(rotated_words: List[Dict]) -> List[Dict]:
    if not rotated_words: return []
    hs = sorted([w["h"] for w in rotated_words if w["h"]>0])
    h_med = hs[len(hs)//2] if hs else 16.0
    y_tol = POST_Y_TOL_FACTOR * h_med

    idxs = list(range(len(rotated_words)))
    idxs.sort(key=lambda i: (rotated_words[i]["cy_rot"], rotated_words[i]["cx_rot"]))
    lines = []
    cur = []

    def flush():
        nonlocal cur
        if not cur: return
        xs = [rotated_words[i]["cx_rot"] for i in cur]
        ys = [rotated_words[i]["cy_rot"] for i in cur]
        m,b = refit_line(list(zip(xs,ys)))
        cur_sorted = sorted(cur, key=lambda i: rotated_words[i]["cx_rot"])
        lines.append(_build_line_result(rotated_words, set(cur_sorted), m, b, rotated=True))
        cur = []

    for i in idxs:
        if not cur:
            cur = [i]
        else:
            y0 = rotated_words[cur[0]]["cy_rot"]
            yi = rotated_words[i]["cy_rot"]
            if abs(yi - y0) <= y_tol:
                cur.append(i)
            else:
                flush()
                cur = [i]
    flush()
    lines.sort(key=lambda L: L["center_y"])
    return lines

# ============================================================
# Utilities: dump lines to txt (only if DEBUG)
# ============================================================
def slope_to_deg(m: float) -> float:
    if math.isinf(m): return 90.0
    return math.degrees(math.atan(m))

def write_lines_txt(base_path: str, suffix: str, lines: List[Dict]) -> Optional[str]:
    if not DEBUG:
        return None
    txt_path = f"{os.path.splitext(base_path)[0]}_{suffix}.txt"
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(f"# {os.path.basename(base_path)}  ({suffix})\n")
        for i, L in enumerate(lines, 1):
            ang = slope_to_deg(L["slope"])
            f.write(f"[{i:03d}] words={L['count']:>3}  slope={ang:+.3f}°\n")
            f.write(L["text"] + "\n\n")
    return txt_path

# ============================================================
# Smart deskew + full pipeline (in-memory; returns words + full_text)
# ============================================================
def smart_deskew_with_lines(image_path: str,
                            out_path: Optional[str] = None,
                            clamp_deg: float = 30.0,
                            use_vision: bool = True) -> Dict:
    img = cv2.imread(image_path, cv2.IMREAD_COLOR)
    if img is None: raise FileNotFoundError(image_path)

    words, full_text = ([], "")
    if use_vision:
        words, full_text = extract_words_and_text(image_path)

    a_h, n_h = estimate_skew_hough(img)
    a_p, n_p = (0.0, 0)
    if words:
        a_p, n_p = estimate_skew_pairs(words, y_band_mult=2.0, min_dx_mult=0.8, max_abs_deg=15.0)

    candidates = []
    if n_h >= 10: candidates += [a_h, -a_h]
    if n_p >= 10: candidates += [a_p, -a_p]
    if not candidates: candidates = [0.0]

    cand = []
    for a in candidates:
        a = float(max(-clamp_deg, min(clamp_deg, a)))
        if all(abs(a - b) > 0.05 for b in cand):
            cand.append(a)

    grid = []
    for a in cand:
        for d in (-0.6, -0.4, -0.2, 0.0, 0.2, 0.4, 0.6):
            g = a + d
            if all(abs(g - x) > 0.05 for x in grid):
                grid.append(g)

    scored = [(a, preview_score(img, -a)) for a in grid]
    best_angle, best_cost = min(scored, key=lambda t: t[1])

    # Debug print kept as a comment
    # print(f"[smart] hough={a_h:.3f}°(n={n_h})  pairs={a_p:.3f}°(n={n_p})  tried={', '.join(f'{a:+.2f}°' for a,_ in scored)}  → chosen {best_angle:+.2f}° (cost={best_cost:.3f})")

    # Rotate in-memory. Save only if DEBUG.
    rotated = rotate_image_keep_bounds(img, -best_angle, border_value=255)
    if DEBUG and out_path:
        cv2.imwrite(out_path, rotated)

    result = {
        "angle_deg": float(best_angle),
        "hough_lines": int(n_h),
        "pair_samples": int(n_p),
        "out_path": out_path if DEBUG else None,
        "pre_txt": None,
        "post_txt": None,
        "pre_lines": [],
        "post_lines": [],
        "words": words,
        "full_text": full_text,
    }

    if words:
        pre_lines = cluster_tilted_lines(words)
        result["pre_lines"] = pre_lines
        result["pre_txt"] = write_lines_txt(image_path, "lines_pre", pre_lines)  # only if DEBUG

        rot_words = transform_words(words, img.shape[:2], -best_angle)
        post_lines = group_horizontal_lines(rot_words)
        result["post_lines"] = post_lines
        result["post_txt"] = write_lines_txt(image_path, "lines_post", post_lines)  # only if DEBUG

        # More debug prints kept as comments
        # def preview(lines, tag):
        #     print(f"  {tag} ({len(lines)} lines)")
        #     for L in lines[:5]:
        #         ang = slope_to_deg(L["slope"])
        #         print(f"    [{L['count']:>3} w] slope={ang:+.3f}° | {L['text'][:90]}")
        # preview(pre_lines, "pre (slope-aware)")
        # preview(post_lines, "post (horizontal)")
        # if DEBUG:
        #     print(f"  → wrote: {result['pre_txt']}  and  {result['post_txt']}")

    return result

# ============================================================
# Multi-client extraction from post lines (robust)
# ============================================================
MEMBER_RE   = re.compile(r'\bMEMBER NAME\s*:\s*(.+)', re.IGNORECASE)
MEMBERID_RE = re.compile(r'\bMEMBER ID\s*:\s*([A-Za-z0-9]+)', re.IGNORECASE)
ICN_LINE_RE = re.compile(r'^\s*\d{12,}\b')

AMOUNT_RE   = re.compile(r'(\d{1,3}(?:,\d{3})*\.\d{2})')  # decimals only
DATE6_RE    = re.compile(r'\b\d{6}\b')
PD_ROW_RE   = re.compile(r'\bPD\s+(D?\d{4})\b', re.IGNORECASE)
TOOTH_RE  = re.compile(r'^(?:[1-9]|[12][0-9]|3[0-2]|[A-Ta-t])$')
SURFACE_RE = re.compile(r'^[MDBOILFP]{1,4}$', re.IGNORECASE)

def _to_float(s: str) -> float:
    try:
        return float(s.replace(',', ''))
    except Exception:
        return 0.0

def _parse_pd_line(t: str) -> Optional[Tuple[str, Optional[float], Optional[float], Optional[float], Optional[str], Optional[str], Optional[str]]]:
    """
    Parse a single PD line.
    Returns: (CDT, billed, allowed, paid, date6, tooth, surface)
    """
    m = PD_ROW_RE.search(t)
    if not m:
        return None

    code = m.group(1)
    code = code if code.upper().startswith('D') else f'D{code}'

    amts = [_to_float(x) for x in AMOUNT_RE.findall(t)]
    billed = allowed = paid = None
    if len(amts) >= 3:
        billed, allowed, paid = amts[-3:]

    d = None
    md = DATE6_RE.search(t)
    if md:
        d = md.group(0)

    tooth = None
    surface = None

    tokens = t.split()
    try:
        code_idx = tokens.index(code)
    except ValueError:
        code_idx = None
        for i, tok in enumerate(tokens):
            if PD_ROW_RE.match(f'PD {tok}'):
                code_idx = i
                break

    if code_idx is not None:
        date_idx = None
        for i in range(code_idx + 1, len(tokens)):
            if DATE6_RE.fullmatch(tokens[i]):
                date_idx = i
                break

        window = tokens[code_idx + 1: date_idx if date_idx is not None else len(tokens)]

        for tok in window:
            if TOOTH_RE.fullmatch(tok):
                tooth = tok.upper()
                break

        start_j = 0
        if tooth is not None:
            for j, tok in enumerate(window):
                if tok.upper() == tooth:
                    start_j = j + 1
                    break
        for tok in window[start_j:]:
            if SURFACE_RE.fullmatch(tok):
                surface = tok.upper()
                break

    return code, billed, allowed, paid, d, tooth, surface

def extract_all_clients_from_lines(post_lines: List[dict]) -> List[dict]:
    """
    Split strictly by MEMBER NAME lines; ignore anything before the first name.
    For each member block, look up ICN from the nearest line above the member header.
    Parse each PD line for CDT, Date SVC, Billed, Allowed, Paid (decimals only).
    """
    texts = [L["text"] for L in post_lines]
    starts = [i for i,t in enumerate(texts) if MEMBER_RE.search(t)]
    if not starts:
        return []

    out_rows = []

    for si, start in enumerate(starts):
        end = starts[si+1] if si+1 < len(starts) else len(texts)

        # header line with MEMBER NAME
        name_line = texts[start]
        raw_name = MEMBER_RE.search(name_line).group(1).strip()
        # Stop at "MEMBER ID" (case-insensitive) and other headers
        cut_points = ["MEMBER ID", "OTH INS CD", "PA:", "DIAG:"]
        mname = raw_name
        for cp in cut_points:
            idx = mname.upper().find(cp)
            if idx != -1:
                mname = mname[:idx].strip()
        # Debug
        # print(raw_name); print(mname)

        # member id: search within the block
        mid = ""
        for t in texts[start:end]:
            m = MEMBERID_RE.search(t)
            if m:
                mid = m.group(1).strip()
                break

        # ICN: search a few lines ABOVE the member header
        icn = ""
        for k in range(start-1, max(-1, start-6), -1):
            if k < 0: break
            mm = ICN_LINE_RE.match(texts[k])
            if mm:
                icn = mm.group(0)
                break

        # PD lines in the block
        had_pd = False
        for t in texts[start:end]:
            if " PD " not in f" {t} ":
                continue
            parsed = _parse_pd_line(t)
            if not parsed:
                continue
            had_pd = True
            code, billed, allowed, paid, dsvc, tooth, surface = parsed
            out_rows.append({
                'Patient Name': mname.title() if mname else "",
                'Patient ID': mid,
                'ICN': icn,
                'CDT Code': code,
                'Tooth': tooth if tooth else "",
                #'Surface': surface if surface else "",
                'Date SVC': dsvc if dsvc else "",
                'Billed Amount': billed if billed is not None else "",
                'Allowed Amount': allowed if allowed is not None else "",
                'Paid Amount':   paid   if paid   is not None else "",
                'Extraction Success': True,
            })

        if not had_pd:
            out_rows.append({
                'Patient Name': mname.title() if mname else "",
                'Patient ID': mid,
                'ICN': icn,
                'CDT Code': "",
                'Tooth': "",
                #'Surface': "",
                'Date SVC': "",
                'Billed Amount': "",
                'Allowed Amount': "",
                'Paid Amount': "",
                'Extraction Success': bool(mname or mid),
            })

    return out_rows

# ============================================================
# ExcelGenerator
# ============================================================
class ExcelGenerator:
    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')

    def create_excel_file(self, df: pd.DataFrame) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Medical Billing Extract"
        ws['A1'] = f"Medical Billing OCR Extract - Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = self.center_alignment
        ws.append([])

        excel_df = self.prepare_dataframe_for_excel(df)
        for r in dataframe_to_rows(excel_df, index=False, header=True):
            ws.append(r)

        self.format_worksheet(ws, len(excel_df) + 3)
        self.add_summary_sheet(wb, excel_df)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def prepare_dataframe_for_excel(self, df: pd.DataFrame) -> pd.DataFrame:
        excel_df = df.copy()
        column_order = [
            'Patient Name', 'Patient ID', 'ICN', 'CDT Code', 'Tooth', 'Date SVC', #'Surface',
            'Billed Amount', 'Allowed Amount', 'Paid Amount',
            'Extraction Success', 'Source File'
        ]
        existing = [c for c in column_order if c in excel_df.columns]
        excel_df = excel_df[existing]
        for amount_col in ['Billed Amount', 'Allowed Amount', 'Paid Amount']:
            if amount_col in excel_df.columns:
                excel_df[amount_col] = excel_df[amount_col].apply(self.format_currency)
        if 'Extraction Success' in excel_df.columns:
            excel_df['Extraction Success'] = excel_df['Extraction Success'].apply(lambda x: 'Yes' if x else 'No')
        return excel_df

    def format_currency(self, value):
        if pd.isna(value) or value == "":
            return ""
        try:
            if isinstance(value, str):
                clean_value = value.replace('$', '').replace(',', '')
                value = float(clean_value)
            return f"${value:,.2f}"
        except (ValueError, TypeError):
            return str(value)

    def format_worksheet(self, ws, data_rows):
        header_row = 3
        for cell in ws[header_row]:
            if cell.value:
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border
        for row in range(header_row + 1, data_rows + 1):
            for cell in ws[row]:
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        self.auto_adjust_columns(ws)
        self.add_conditional_formatting(ws, header_row, data_rows)

    def auto_adjust_columns(self, ws):
        max_col = ws.max_column
        max_row = ws.max_row
        for col_idx in range(1, max_col + 1):
            max_len = 0
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue
                try:
                    val = cell.value
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
                except Exception:
                    pass
            letter = get_column_letter(col_idx)
            ws.column_dimensions[letter].width = min(max_len + 2, 50)

    def add_conditional_formatting(self, ws, header_row, data_rows):
        success_col = None
        for col, cell in enumerate(ws[header_row], 1):
            if cell.value == 'Extraction Success':
                success_col = col
                break
        if success_col:
            for row in range(header_row + 1, data_rows + 1):
                cell = ws.cell(row=row, column=success_col)
                if cell.value == 'Yes':
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif cell.value == 'No':
                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

    def add_summary_sheet(self, wb, df):
        ws = wb.create_sheet(title="Summary")
        ws['A1'] = "Extraction Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        row = 3
        stats = [
            ("Total Rows", len(df)),
            ("Successful", len(df[df['Extraction Success'] == 'Yes']) if 'Extraction Success' in df.columns else 0),
            ("Failed", len(df[df['Extraction Success'] == 'No']) if 'Extraction Success' in df.columns else 0),
        ]
        for name, val in stats:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = val
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        ExcelGenerator().auto_adjust_columns(ws)
        row += 2
        ws[f'A{row}'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{row}'].font = Font(italic=True)

# ============================================================
# Runner: glue everything together
# ============================================================
def process_images_to_excel(files: List[str], out_excel: str, deskewed_only: bool=False) -> None:
    excel_gen = ExcelGenerator()
    records: List[Dict[str, Any]] = []

    for src in files:
        try:
            if deskewed_only:
                img = cv2.imread(src, cv2.IMREAD_COLOR)
                if img is None:
                    raise FileNotFoundError(src)
                words, _ = extract_words_and_text(src)
                rot_words = []
                for w in words:
                    ww = dict(w)
                    ww["cx_rot"], ww["cy_rot"] = w["cx"], w["cy"]
                    rot_words.append(ww)
                post_lines = group_horizontal_lines(rot_words)

                post_txt = write_lines_txt(src, "lines_post", post_lines)  # only if DEBUG

                rows = extract_all_clients_from_lines(post_lines)
                for r in rows:
                    r["Source File"] = os.path.basename(src)
                    records.append(r)
                # if DEBUG: print(f"{src} → parsed {len(rows)} PD rows (wrote {post_txt})")

            else:
                base, ext = os.path.splitext(src)
                dst = f"{base}_deskewed{ext if ext else '.jpg'}" if DEBUG else None
                info = smart_deskew_with_lines(src, dst, clamp_deg=30.0, use_vision=True)
                post_lines = info.get("post_lines", []) if info else []
                rows = extract_all_clients_from_lines(post_lines) if post_lines else []
                for r in rows:
                    r["Source File"] = os.path.basename(src)
                    records.append(r)
                # if DEBUG: print(f"{src} → rotated by {-info['angle_deg']:.3f}° → {dst}")

        except Exception as e:
            # if DEBUG: print(f"{src}: {e}")
            records.append({
                'Patient Name': "", 'Patient ID': "", 'ICN': "", 'CDT Code': "",
                'Date SVC': "", 'Billed Amount': "", 'Allowed Amount': "", 'Paid Amount': "",
                'Extraction Success': False, 'Source File': os.path.basename(src),
            })

    df = pd.DataFrame.from_records(records)
    data = excel_gen.create_excel_file(df)
    with open(out_excel, "wb") as f:
        f.write(data)
    # if DEBUG:
    #     print(f"\n✅ Wrote Excel → {out_excel}")
    #     print("   (and per-image: *_lines_pre.txt, *_lines_post.txt, *_deskewed.* when DEBUG=True)")

# ============================================================
# CLI
# ============================================================
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", help="Folder of images (jpg/png/tif).", default=None)
    ap.add_argument("--files", nargs="*", help="Specific image files.", default=None)
    ap.add_argument("--out", help="Output Excel path.", required=True)
    ap.add_argument("--deskewed-only", action="store_true",
                    help="Only process files whose name contains '_deskewed'; skip deskew step.")
    args = ap.parse_args()

    paths: List[str] = []
    if args.files:
        for f in args.files:
            if os.path.isfile(f):
                paths.append(f)
    if args.input and os.path.isdir(args.input):
        for ext in ("*.jpg","*.jpeg","*.png","*.tif","*.tiff","*.bmp"):
            paths.extend(glob.glob(os.path.join(args.input, ext)))

    if args.deskewed_only:
        paths = [p for p in paths if "_deskewed" in os.path.basename(p).lower()]

    if not paths:
        raise SystemExit("No input images found. Use --files or --input (and --deskewed-only if desired).")

    if not os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"):
        # print("WARNING: GOOGLE_APPLICATION_CREDENTIALS not set. Set it to your local service account JSON path.")
        pass

    process_images_to_excel(paths, args.out, deskewed_only=args.deskewed_only)

if __name__ == "__main__":
    main()
